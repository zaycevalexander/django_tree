from typing import List

import openpyxl

from .base_reader import BaseReader


class ExcelReader(BaseReader):

    def __init__(self, file):
        super().__init__(file)
        self.wb: openpyxl.Workbook = openpyxl.load_workbook(file)
        self.sheet = self.wb.active

    @property
    def items(self) -> List:
        for i in range(1, self.sheet.max_row + 1):
            yield {'email': self.sheet.cell(i, 1).value}
